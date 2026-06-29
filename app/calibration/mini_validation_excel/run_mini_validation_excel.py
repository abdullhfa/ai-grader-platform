"""
Build synthetic .xlsx fixtures and run project_profile + evidence_layer.

Usage (repo root):
  python -m app.calibration.mini_validation_excel.run_mini_validation_excel

Observation-only — fill observed_friction / unexpected_result after human review.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional


def _build_workbook(case_id: str, dest: Path) -> Optional[str]:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return "openpyxl_not_installed"

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    if case_id == "case_a_full_workbook":
        ws = wb.active
        ws.title = "Sales"
        ws["A1"], ws["A2"], ws["A3"] = 100, 200, 50
        ws["B1"] = "=SUM(A1:A3)"
        ws["B2"] = "=AVERAGE(A1:A3)"
        ws["B3"] = '=IF(B1>300,"High","Low")'
        ws["C1"], ws["C2"], ws["C3"] = "A", "B", "C"
        ws["D1"] = "=VLOOKUP(C1,A1:B3,2,FALSE)"
        ws2 = wb.create_sheet("Charts")
        ws2["A1"], ws2["A2"] = 10, 20
        ws2["B1"] = "=Sales!B1"
        tab = Table(displayName="SalesTable", ref="A1:B3")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        chart = BarChart()
        chart.title = "Sales Chart"
        data = Reference(ws, min_col=1, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=False)
        ws.add_chart(chart, "F2")
        wb.save(dest)
        wb.close()
        return None

    if case_id == "case_b_values_only":
        ws = wb.active
        ws.title = "Data"
        ws["A1"], ws["A2"], ws["A3"] = 10, 20, 30
        ws["B1"], ws["B2"] = "Revenue", "Cost"
        wb.save(dest)
        wb.close()
        return None

    if case_id == "case_c_formula_errors":
        ws = wb.active
        ws.title = "Errors"
        ws["A1"] = "=1/0"
        ws["A2"] = "=#REF!+1"
        ws["A3"] = "=SUM(#REF!)"
        wb.save(dest)
        wb.close()
        return None

    if case_id == "case_d_charts_only":
        ws = wb.active
        ws.title = "ChartSheet"
        ws["A1"], ws["A2"], ws["A3"] = 5, 10, 15
        chart = BarChart()
        data = Reference(ws, min_col=1, min_row=1, max_row=3)
        chart.add_data(data, titles_from_data=False)
        ws.add_chart(chart, "C2")
        wb.save(dest)
        wb.close()
        return None

    if case_id == "case_e_cross_sheet_logic":
        data = wb.active
        data.title = "Data"
        data["A1"], data["A2"], data["A3"] = 5, 12, 0
        data["B1"], data["B2"] = 100, 50
        calc = wb.create_sheet("Calc")
        calc["A1"] = "=Data!A1+Data!A2"
        calc["A2"] = '=IF(Data!A3>0,Data!B1,Data!B2)'
        calc["A3"] = "=COUNTIF(Data!A1:A3,\">0\")"
        calc["A4"] = "=SUMIF(Data!A1:A3,\">0\",Data!B1:B3)"
        wb.save(dest)
        wb.close()
        return None

    wb.close()
    return f"unknown_case:{case_id}"


def _bootstrap_case(case_id: str, case_dir: Path) -> None:
    xlsx = case_dir / "workbook.xlsx"
    err = _build_workbook(case_id, xlsx)
    if err:
        raise RuntimeError(err)


def _collect_files(case_dir: Path) -> List[str]:
    return [str(p.resolve()) for p in sorted(case_dir.rglob("*")) if p.is_file()]


def _compact(profile: Dict[str, Any], evidence_layer: Dict[str, Any]) -> Dict[str, Any]:
    xl = profile.get("excel_semantic_evidence") or {}
    agg = xl.get("aggregate") or {}
    xl_items = [
        it
        for it in (evidence_layer.get("items") or [])
        if isinstance(it, dict) and it.get("evidence_type") == "spreadsheet_semantic"
    ]
    return {
        "engines_detected": profile.get("engines_detected"),
        "excel_aggregate": agg,
        "spreadsheet_semantic_summary": xl.get("spreadsheet_semantic_summary"),
        "excel_noise_flags": xl.get("noise_flags"),
        "extractions_head": (xl.get("extractions") or [])[:1],
        "evidence_layer_xl_items": [
            {
                "evidence_type": it.get("evidence_type"),
                "formula_cells": it.get("formula_cells"),
                "formula_types": it.get("formula_types"),
                "charts_detected": it.get("charts_detected"),
                "chart_count": it.get("chart_count"),
                "cross_sheet_references": it.get("cross_sheet_references"),
                "conditional_logic_detected": it.get("conditional_logic_detected"),
                "execution_evidence": it.get("execution_evidence"),
            }
            for it in xl_items
        ],
    }


def run_all() -> Dict[str, Any]:
    root = Path(__file__).resolve().parent
    template = json.loads((root / "expected_cases.json").read_text(encoding="utf-8"))

    from app.project_intelligence.evidence_schema import build_evidence_layer_from_profile
    from app.project_intelligence.project_profile import build_project_profile

    cases_out: List[Dict[str, Any]] = []
    for entry in template.get("cases") or []:
        cid = entry.get("case_id") or ""
        case_dir = root / "cases" / cid
        case_dir.mkdir(parents=True, exist_ok=True)
        _bootstrap_case(cid, case_dir)
        paths = _collect_files(case_dir)
        profile = build_project_profile(paths)
        evidence_layer = build_evidence_layer_from_profile(profile)
        snap = _compact(profile, evidence_layer)
        cases_out.append(
            {
                **entry,
                "actual_behavior": json.dumps(snap, ensure_ascii=False, indent=2),
            }
        )

    return {
        "run_purpose": template.get("run_purpose", ""),
        "cases": cases_out,
        "note": "Fill observed_friction and unexpected_result after manual review.",
    }


def main() -> None:
    out = run_all()
    root = Path(__file__).resolve().parent
    out_path = root / "mini_validation_excel_last_run.json"
    if out_path.is_file():
        prev = json.loads(out_path.read_text(encoding="utf-8"))
        prev_by = {c.get("case_id"): c for c in prev.get("cases") or []}
        for row in out.get("cases") or []:
            cid = row.get("case_id")
            if cid in prev_by:
                for key in ("observed_friction", "unexpected_result", "observation_notes"):
                    if prev_by[cid].get(key):
                        row[key] = prev_by[cid][key]
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
