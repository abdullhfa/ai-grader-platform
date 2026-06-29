"""
Deterministic Excel / spreadsheet semantics — formulas, charts, tables (no AI grading).

Uses openpyxl with data_only=False so formula strings are preserved, not cached values.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

EXCEL_SEMANTIC_EXTRACTOR_VERSION = "1.0"
EXCEL_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
LEGACY_XLS_EXT = ".xls"
MAX_WORKBOOK_FILES = 5
MAX_SHEETS_SCAN = 32
MAX_ROWS_SCAN = 600
MAX_COLS_SCAN = 80
MAX_FORMULA_SAMPLE = 120

_CONDITIONAL_FUNCS = frozenset(
    {
        "IF",
        "IFS",
        "IFERROR",
        "IFNA",
        "COUNTIF",
        "SUMIF",
        "AVERAGEIF",
        "COUNTIFS",
        "SUMIFS",
        "AVERAGEIFS",
        "MAXIFS",
        "MINIFS",
    }
)

_FORMULA_FUNC_RE = re.compile(
    r"(?<![A-Z0-9.])"
    r"(SUM|SUMIF|SUMIFS|AVERAGE|AVERAGEIF|AVERAGEIFS|IF|IFS|IFERROR|IFNA|"
    r"VLOOKUP|HLOOKUP|XLOOKUP|INDEX|MATCH|COUNT|COUNTA|COUNTIF|COUNTIFS|"
    r"MIN|MAX|MINIFS|MAXIFS|ROUND|ROUNDDOWN|ROUNDUP|AND|OR|NOT|"
    r"OFFSET|INDIRECT|TEXT|CONCATENATE|CONCAT|LEN|LEFT|RIGHT|MID|"
    r"DATE|DATEDIF|YEAR|MONTH|DAY|TODAY|NOW|PMT|FV|NPV|IRR)"
    r"(?=\s*\()",
    re.IGNORECASE,
)

_CROSS_SHEET_REF_RE = re.compile(
    r"(?:'([^']{1,120})'|([A-Za-z_][\w\s\.]{0,80}))!\$?[A-Za-z]{1,4}\$?\d+",
    re.IGNORECASE,
)

_FORMULA_ERROR_TOKEN_RE = re.compile(
    r"#(?:REF!|DIV/0!|N/A|VALUE!|NAME\?|NULL!|NUM!)",
    re.IGNORECASE,
)
_DIV_ZERO_FORMULA_RE = re.compile(r"/\s*0(?:\)|$|[,;\s])")

def _try_load_workbook(path: Path) -> Tuple[Any, Optional[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None, "openpyxl_not_installed"
    try:
        wb = load_workbook(path, read_only=False, data_only=False, keep_links=False)
        return wb, None
    except Exception as exc:  # noqa: BLE001
        return None, f"workbook_load_failed:{exc.__class__.__name__}"


def _cell_formula_text(cell: Any) -> Optional[str]:
    val = getattr(cell, "value", None)
    if val is None:
        return None
    if getattr(cell, "data_type", None) == "f":
        return str(val) if val is not None else None
    if isinstance(val, str) and val.strip().startswith("="):
        return val.strip()
    return None


def _extract_formula_functions(formula: str) -> Set[str]:
    found: Set[str] = set()
    for m in _FORMULA_FUNC_RE.finditer(formula):
        found.add(m.group(1).upper())
    return found


def _cross_sheet_refs_in_formula(
    formula: str, sheet_names: Set[str], current_sheet: str
) -> bool:
    cur = (current_sheet or "").strip().lower()
    for m in _CROSS_SHEET_REF_RE.finditer(formula):
        ref_sheet = (m.group(1) or m.group(2) or "").strip()
        if not ref_sheet or ref_sheet not in sheet_names:
            continue
        if ref_sheet.strip().lower() != cur:
            return True
    return False


def _formula_has_error_signal(formula: str) -> bool:
    if _FORMULA_ERROR_TOKEN_RE.search(formula):
        return True
    return bool(_DIV_ZERO_FORMULA_RE.search(formula))


def _scan_worksheet(ws: Any, sheet_names: Set[str], current_sheet: str) -> Dict[str, Any]:
    formula_cells = 0
    value_cells = 0
    formula_types: Set[str] = set()
    conditional_logic = False
    cross_sheet = False
    formula_errors = 0
    formula_samples: List[str] = []

    max_row = min(int(getattr(ws, "max_row", None) or 1), MAX_ROWS_SCAN)
    max_col = min(int(getattr(ws, "max_column", None) or 1), MAX_COLS_SCAN)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            formula = _cell_formula_text(cell)
            if formula:
                formula_cells += 1
                if len(formula_samples) < 8:
                    formula_samples.append(formula[:200])
                funcs = _extract_formula_functions(formula)
                formula_types.update(funcs)
                if funcs & _CONDITIONAL_FUNCS:
                    conditional_logic = True
                if _cross_sheet_refs_in_formula(formula, sheet_names, current_sheet):
                    cross_sheet = True
                if _formula_has_error_signal(formula):
                    formula_errors += 1
            elif getattr(cell, "value", None) is not None:
                value_cells += 1

    chart_count = 0
    charts_detected = False
    try:
        charts = getattr(ws, "_charts", None) or []
        chart_count = len(charts)
        charts_detected = chart_count > 0
    except Exception:
        charts_detected = False

    table_count = 0
    structured_tables: List[str] = []
    try:
        tables_obj = getattr(ws, "tables", None)
        if tables_obj:
            if hasattr(tables_obj, "values"):
                for tbl in tables_obj.values():
                    name = getattr(tbl, "displayName", None) or getattr(tbl, "name", None)
                    ref = getattr(tbl, "ref", None)
                    structured_tables.append(
                        str(name or ref or "table")
                    )
            elif isinstance(tables_obj, dict):
                for tbl in tables_obj.values():
                    name = getattr(tbl, "displayName", None) or getattr(tbl, "name", None)
                    structured_tables.append(str(name or "table"))
        table_count = len(structured_tables)
    except Exception:
        table_count = 0

    noise: List[Dict[str, str]] = []
    if value_cells > 0 and formula_cells == 0:
        noise.append({"flag": "values_without_formulas"})
    if formula_errors > 0:
        noise.append({"flag": "formula_errors_detected"})
    if charts_detected and formula_cells == 0:
        noise.append({"flag": "charts_without_formulas"})

    return {
        "sheet_name": str(getattr(ws, "title", "") or ""),
        "formula_cells": formula_cells,
        "value_cells": value_cells,
        "formula_types": sorted(formula_types),
        "charts_detected": charts_detected,
        "chart_count": chart_count,
        "table_count": table_count,
        "structured_table_names": structured_tables[:10],
        "cross_sheet_references": cross_sheet,
        "conditional_logic_detected": conditional_logic,
        "formula_error_cells": formula_errors,
        "noise_flags": noise,
        "formula_samples": formula_samples,
    }


def extract_single_workbook(path: Path) -> Dict[str, Any]:
    wb, err = _try_load_workbook(path)
    row: Dict[str, Any] = {
        "path": str(path),
        "basename": path.name,
        "readable": wb is not None,
        "load_error": err,
    }
    if wb is None:
        row["signals"] = None
        flags = [{"flag": "workbook_unreadable"}]
        if err == "openpyxl_not_installed":
            flags = [{"flag": "openpyxl_not_installed"}]
        row["noise_flags"] = flags
        return row

    try:
        sheet_names_list = list(wb.sheetnames or [])
        sheet_names = set(sheet_names_list)
        sheets_out: List[Dict[str, Any]] = []
        total_formulas = 0
        all_formula_types: Set[str] = set()
        chart_count = 0
        table_count = 0
        cross_sheet = False
        conditional = False
        formula_errors = 0
        all_noise: List[Dict[str, str]] = []

        for name in sheet_names_list[:MAX_SHEETS_SCAN]:
            ws = wb[name]
            sh = _scan_worksheet(ws, sheet_names, str(name))
            sheets_out.append(sh)
            total_formulas += int(sh.get("formula_cells") or 0)
            all_formula_types.update(sh.get("formula_types") or [])
            chart_count += int(sh.get("chart_count") or 0)
            table_count += int(sh.get("table_count") or 0)
            cross_sheet = cross_sheet or bool(sh.get("cross_sheet_references"))
            conditional = conditional or bool(sh.get("conditional_logic_detected"))
            formula_errors += int(sh.get("formula_error_cells") or 0)
            all_noise.extend(sh.get("noise_flags") or [])

        signals = {
            "sheet_count": len(sheet_names_list),
            "named_sheets": sheet_names_list[:MAX_SHEETS_SCAN],
            "formula_cells": total_formulas,
            "formula_types": sorted(all_formula_types),
            "charts_detected": chart_count > 0,
            "chart_count": chart_count,
            "table_count": table_count,
            "structured_tables_detected": table_count > 0,
            "cross_sheet_references": cross_sheet,
            "conditional_logic_detected": conditional,
            "formula_error_cells": formula_errors,
        }

        if total_formulas == 0:
            all_noise.append({"flag": "values_without_formulas"})
        if formula_errors > 0:
            all_noise.append({"flag": "formula_errors_detected"})
        if chart_count > 0 and total_formulas == 0:
            all_noise.append({"flag": "charts_without_formulas"})

        seen_nf: Set[str] = set()
        deduped_nf: List[Dict[str, str]] = []
        for nf in all_noise:
            f = nf.get("flag")
            if not f or f in seen_nf:
                continue
            seen_nf.add(f)
            deduped_nf.append({"flag": f})

        row["signals"] = signals
        row["sheets"] = sheets_out
        row["noise_flags"] = deduped_nf
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return row


def list_excel_workbook_files(file_paths: Sequence[Path]) -> Tuple[List[Path], List[Path]]:
    """Return (openpyxl_supported, legacy_xls_only)."""
    supported: List[Path] = []
    legacy: List[Path] = []
    for fp in sorted(file_paths, key=lambda p: str(p).lower()):
        try:
            suf = fp.suffix.lower()
            if suf in EXCEL_EXTENSIONS and fp.is_file():
                supported.append(fp.resolve())
            elif suf == LEGACY_XLS_EXT and fp.is_file():
                legacy.append(fp.resolve())
        except OSError:
            continue
    return supported[:MAX_WORKBOOK_FILES], legacy


def _merge_extractions(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    readable = [r for r in rows if r.get("readable") and isinstance(r.get("signals"), dict)]
    if not readable:
        return {
            "sheet_count": 0,
            "named_sheets": [],
            "formula_cells": 0,
            "formula_types": [],
            "charts_detected": False,
            "chart_count": 0,
            "table_count": 0,
            "structured_tables_detected": False,
            "cross_sheet_references": False,
            "conditional_logic_detected": False,
            "formula_error_cells": 0,
        }

    def _sum_int(key: str) -> int:
        return sum(int((r.get("signals") or {}).get(key) or 0) for r in readable)

    def _max_int(key: str) -> int:
        return max(int((r.get("signals") or {}).get(key) or 0) for r in readable)

    def _any_bool(key: str) -> bool:
        return any(bool((r.get("signals") or {}).get(key)) for r in readable)

    types: Set[str] = set()
    sheets: List[str] = []
    for r in readable:
        sig = r.get("signals") or {}
        types.update(sig.get("formula_types") or [])
        for n in sig.get("named_sheets") or []:
            if n and n not in sheets:
                sheets.append(str(n))

    return {
        "sheet_count": _max_int("sheet_count"),
        "named_sheets": sheets[:MAX_SHEETS_SCAN],
        "formula_cells": _sum_int("formula_cells"),
        "formula_types": sorted(types),
        "charts_detected": _any_bool("charts_detected"),
        "chart_count": _sum_int("chart_count"),
        "table_count": _sum_int("table_count"),
        "structured_tables_detected": _any_bool("structured_tables_detected"),
        "cross_sheet_references": _any_bool("cross_sheet_references"),
        "conditional_logic_detected": _any_bool("conditional_logic_detected"),
        "formula_error_cells": _sum_int("formula_error_cells"),
    }


def build_spreadsheet_semantic_summary(block: Dict[str, Any]) -> Dict[str, Any]:
    """Academic sufficiency slice — presence of spreadsheet semantics only."""
    agg = block.get("aggregate") or {}
    files = block.get("workbook_files") or []
    noise: List[Dict[str, str]] = list(block.get("noise_flags") or [])
    for row in block.get("extractions") or []:
        if not isinstance(row, dict):
            continue
        for nf in row.get("noise_flags") or []:
            if isinstance(nf, dict) and nf.get("flag"):
                noise.append(nf)

    seen: Set[str] = set()
    deduped: List[Dict[str, str]] = []
    for nf in noise:
        f = nf.get("flag")
        if not f or f in seen:
            continue
        seen.add(f)
        deduped.append({"flag": f})

    return {
        "workbook_file_count": len(files),
        "readable_workbook_count": sum(
            1 for r in (block.get("extractions") or []) if isinstance(r, dict) and r.get("readable")
        ),
        "workbook": {
            "sheet_count": agg.get("sheet_count"),
            "named_sheets": agg.get("named_sheets") or [],
        },
        "formulas": {
            "formula_cells": agg.get("formula_cells"),
            "formula_types": agg.get("formula_types") or [],
            "conditional_logic_detected": agg.get("conditional_logic_detected"),
            "cross_sheet_references": agg.get("cross_sheet_references"),
            "formula_error_cells": agg.get("formula_error_cells"),
        },
        "charts": {
            "charts_detected": agg.get("charts_detected"),
            "chart_count": agg.get("chart_count"),
        },
        "tables": {
            "table_count": agg.get("table_count"),
            "structured_tables_detected": agg.get("structured_tables_detected"),
        },
        "spreadsheet_noise_flags": sorted(deduped, key=lambda x: x["flag"]),
        "limitations_ar": (
            "استخراج حتمي من بنية ملف Excel (صيغ، مخططات، جداول) عبر openpyxl مع data_only=False؛ "
            "لا يقيّم صحة الحسابات أو منطق الأعمال؛ ملفات .xls القديمة غير مدعومة."
        ),
    }


def build_spreadsheet_semantic_evidence_items(
    block: Dict[str, Any],
    engines: List[str],
) -> List[Dict[str, Any]]:
    engine = "excel_spreadsheet" if "excel_spreadsheet" in engines else (
        engines[0] if engines else "unknown"
    )
    agg = block.get("aggregate") or {}
    items: List[Dict[str, Any]] = []
    for row in block.get("extractions") or []:
        if not isinstance(row, dict) or not row.get("readable"):
            continue
        sig = row.get("signals") or {}
        items.append(
            {
                "evidence_type": "spreadsheet_semantic",
                "engine": engine,
                "system": None,
                "confidence": None,
                "execution_evidence": "excel_workbook",
                "evidence_count": sig.get("formula_cells"),
                "sources": [str(row.get("path") or "")],
                "basename": row.get("basename"),
                "sheet_count": sig.get("sheet_count"),
                "formula_cells": sig.get("formula_cells"),
                "formula_types": sig.get("formula_types") or [],
                "charts_detected": sig.get("charts_detected"),
                "chart_count": sig.get("chart_count"),
                "table_count": sig.get("table_count"),
                "cross_sheet_references": sig.get("cross_sheet_references"),
                "conditional_logic_detected": sig.get("conditional_logic_detected"),
                "criterion_candidates": [],
                "weight": 0.0,
            }
        )
    if not items and block.get("workbook_files"):
        items.append(
            {
                "evidence_type": "spreadsheet_semantic",
                "engine": engine,
                "system": None,
                "confidence": None,
                "execution_evidence": "excel_workbook_unreadable",
                "evidence_count": 0,
                "sources": list(block.get("workbook_files") or [])[:MAX_WORKBOOK_FILES],
                "formula_cells": 0,
                "formula_types": [],
                "charts_detected": False,
                "chart_count": 0,
                "cross_sheet_references": False,
                "conditional_logic_detected": False,
                "criterion_candidates": [],
                "weight": 0.0,
            }
        )
    if items and agg:
        items[0]["aggregate_snapshot"] = {
            "formula_types": agg.get("formula_types"),
            "sheet_count": agg.get("sheet_count"),
        }
    return items


def extract_excel_semantic_evidence(file_paths: Sequence[Path]) -> Dict[str, Any]:
    supported, legacy = list_excel_workbook_files(file_paths)
    extractions = [extract_single_workbook(p) for p in supported]
    aggregate = _merge_extractions(extractions)

    noise_flags: List[Dict[str, str]] = []
    if legacy:
        noise_flags.append({"flag": "legacy_xls_not_supported"})
    if supported and not any(r.get("readable") for r in extractions):
        noise_flags.append({"flag": "all_workbooks_unreadable"})
    if not supported and not legacy:
        noise_flags.append({"flag": "no_excel_workbooks_found"})

    return {
        "version": 1,
        "extractor_version": EXCEL_SEMANTIC_EXTRACTOR_VERSION,
        "workbook_files": [str(p) for p in supported],
        "legacy_xls_files": [str(p) for p in legacy[:5]],
        "extractions": extractions,
        "aggregate": aggregate,
        "spreadsheet_semantic_summary": {},
        "noise_flags": noise_flags,
        "notes_ar": (
            "تحليل Excel حتمي: صيغ ومخططات وجداول؛ "
            "لا يستبدل تقييم المعلم ولا يفسر منطق الأعمال."
        ),
    }


def finalize_excel_semantic_block(block: Dict[str, Any]) -> Dict[str, Any]:
    if not block:
        return block
    block = dict(block)
    block["spreadsheet_semantic_summary"] = build_spreadsheet_semantic_summary(block)
    return block
